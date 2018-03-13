#! /usr/bin/python
import openpyxl as opExcel
import numpy as np
from py2neo import Graph
import sys
import time

class Student(object):
    name = ""
    surname = ""
    pos = ""
    isTutor = False
    
    def __init__(self, name, surname, pos, isTutor):
        self.name = name
        self.surname = surname
        self.pos = pos
        self.isTutor = isTutor

class Question(object):
    name = ""
    pos = ""

    def __init__(self, name, pos):
        self.name = name
        self.pos = pos

def str2bool(string):
    return str(string).lower() in ("yes", "true", "t", "1", "=true()")

wb = opExcel.load_workbook(filename = 'questions.xlsx')
ws = wb.active

#Variables
questionArray = []
studentsArray = []

#Get the header (title for the questions)
for rows in ws.iter_rows(min_col=4, max_col=ws.max_column, max_row=1):
    for cell in rows:
        pos = cell.column
        name = cell.value
        questionArray.append(Question(name, pos))        

#Get the students info
for rows in ws.iter_rows(max_col=3, min_row=2, max_row=ws.max_row):
    for cell in rows:
        if(cell.column == 'A'):
            name = cell.value
        if(cell.column == 'B'):
            surname = cell.value
        if(cell.column == 'C'):
            isTutor = str2bool(cell.value)
            pos = cell.row
    studentsArray.append(Student(name, surname, pos, isTutor))



#--------------------------------------------------------------------
# SIMILARITY CALCULATION
#-------------------------------------------------------------------
cur_time = time.clock()
sim_matrix = {}

#Calculate the similarity for each pair of students
for rows1 in ws.iter_rows(min_col=4, max_col=ws.max_column, min_row=2, max_row=5):#ws.max_row-1):   
    user1 = []
    current_row = 0
    for cell1 in rows1:
        user1.append(cell1.value)
        current_row = cell1.row

    for rows2 in ws.iter_rows(min_col=4, max_col=ws.max_column, min_row=current_row+1, max_row=ws.max_row):
        user2 = []
        cr2 = 0
        for cell2 in rows2:
            user2.append(cell2.value)
            cr2 = cell2.row
        
        user1 = np.array(user1)
        user2 = np.array(user2)
        sim = np.linalg.norm(user1-user2)
        sim_matrix[(current_row,cr2)] = sim
        #print 'user %s with user %s --> similarity = %.4f' % (current_row, cr2, sim)

print("Time to calculate similarity:" + str(time.clock() - cur_time))


#-------------------------------------------------------------------
# SAVE INTO THE GRAPH DB
#-------------------------------------------------------------------

#Connect to the graph
graph = Graph(#bolt = True,
              host = '172.19.0.3',
              #bolt_port = 7687,
              password = 'mmmPassForGraphServer')

def createStudents(students):
    # Open transaction
    tx = graph.begin()
    stmt = "CREATE (n:Person {name:{NAME}, surname:{SURNAME}, isTutor:{isTutor}})"
    for student in students:
        tx.run(stmt, {"NAME": student.name, "SURNAME": student.surname, "isTutor": student.isTutor})
    # Commit transaction
    tx.commit()

def createQuestions(questions):
    tx = graph.begin()
    stmt = "CREATE (n:Question {name:{NAME}})"
    for question in questions:
        tx.run(stmt, {"NAME": question.name})
    tx.commit()

def createStudQuestRel(students, questions):
    stmt = "MATCH (p:Person {name:{PNAME}, surname:{PSURNAME}}), (q:Question {name:{QNAME}}) CREATE (p)-[:ANSWERED {answer:{ANSWER}}]->(q)"
    tx = graph.begin()
    for rows in ws.iter_rows(min_col=4, max_col=ws.max_column, min_row=2, max_row=ws.max_row):
        for cell in rows:
            student = studentsArray[cell.row-2]
            question = questionArray[opExcel.utils.column_index_from_string(cell.column)-4]
            tx.run(stmt, {"PNAME":student.name,
                            "PSURNAME":student.surname,
                            "QNAME":question.name,
                            "ANSWER":cell.value
                            })
    tx.commit()

def createSimilarity(students):
    #stmt = ("MATCH (p1:Person {name:{P1NAME}, surname:{P1SURNAME}}), " 
     #             "(p2:Person {name:{P2NAME}, surname:{P2SURNAME}}) " 
      #            "MERGE (p1)-[:SIMILARITY {euclidian:{ANSWER}}]-(p2)")
	
    stmt = ("MATCH (p1:Person {name:{P1NAME}, surname:{P1SURNAME}})-[x:ANSWERED]->(m:Question)<-[y:ANSWERED]-(p2:Person {isTutor:false})"
	    "WITH  SUM((x.answer - y.answer)^2) AS xyDotProduct,"
            "p1, p2 "
	    "MERGE (p1)-[s:SIMILARITY]-(p2) "
	    "SET s.euclidean = SQRT(xyDotProduct)")
    
    for student in students:
        tx = graph.begin()
        sys.stdout.write(".")
        if student.isTutor:
            tx.run(stmt, {"P1NAME": student.name,
                          "P1SURNAME": student.surname
                    })
        tx.commit()
            
print("Creating students...")
cur_time = time.clock() 
createStudents(studentsArray)
print(time.clock() - cur_time)

print("Creating question...")
cur_time = time.clock() 
createQuestions(questionArray)
print(time.clock() - cur_time)

print("Creating answers relationship...")
cur_time = time.clock() 
createStudQuestRel(studentsArray,questionArray)
print(time.clock() - cur_time)

print("Creating similarity relationship...")
cur_time = time.clock() 
#createSimilarity(studentsArray)
print(time.clock() - cur_time)

